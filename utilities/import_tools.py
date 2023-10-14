from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

from datetime import datetime, timedelta, date

from ingredients.models import Category, Ingredient, Nutritionals, IngredientImp


def fixVerduraFresca():
    noNut=Ingredient.objects.filter(nutritional__calories=0)
   # print(noNut)
    wb = load_workbook('/home/matte/IdeaProjects/verdure.xlsx')
    ws = wb.active

    ingredient_list=[]
    for row in range(2, 870):
        name = ws["B" + str(row)].value
        cat = ws["H" + str(row)].value
        cal = ws["C" + str(row)].value
        glu = ws["D" + str(row)].value
        pro = ws["E" + str(row)].value
        lip = ws["F" + str(row)].value
        ami = ws["G" + str(row)].value
        if name:
            word=name.partition(' ')[0]
            selection=Ingredient.objects.filter(name__icontains=word).first()
            for nn in noNut:
                if word in nn.name:
                    nn.nutritional.carbohydrates=glu
                    nn.nutritional.fat=lip
                    nn.nutritional.calories=cal
                    nn.nutritional.proteins=pro
                    nn.nutritional.starch=ami
                    nn.save()
                    print('done '+nn.name)

            if selection.nutritional:
                pass
                #print('ok')
            else:
                nut=Nutritionals()
                nut.calories=cal
                nut.carbohydrates=glu
                nut.proteins=pro
                nut.fat=lip
                nut .starch=ami
                nut.save()
                selection.nutritional=nut
                selection.save()



        category= Category.objects.filter(id=1)







"""

def import_suppliers():
    wb = load_workbook('/home/mate/Documenti/supply_supplier.xlsx')
    ws = wb.active
    print(ws)
    for row in range(2, 13):
        name = ws["B" + str(row)].value

        supplier = Supplier()
        supplier.name = name
        print(supplier)
        supplier.save()


def import_articles():
    wb = load_workbook('/home/mate/Documenti/mpoart.xlsx')
    ws = wb.active
    print(ws)
    for row in range(2, 192):
        try:
            name = ws["A" + str(row)].value
            ingredientName= ws["B" + str(row)].value
            supplierName = ws["C" + str(row)].value
            ua = ws["D" + str(row)].value
            supplier= Supplier.objects.filter(name__contains=supplierName).first()
            ingredient=Ingredient.objects.filter(name__contains=ingredientName).first()

            if supplier and ingredient and name:
                print(supplier.name, ingredient.name, name)
                article=Article()
                article.name=name
                article.supplier=supplier
                article.ingredient=ingredient
                article.unitaArrivo=ua
                article.giacenza_in_gr=0
                article.prezzo_unita_arrivo=0

                article.save()
        except:
            print('e' )


def import_arrivi():
    wb = load_workbook('/home/mate/Documenti/mpo.xlsx')
    ws = wb.active
    print(ws)
    c=1
    for row in range(2, 268):

        date = ws["A" + str(row)].value
        articleName= ws["B" + str(row)].value
        supplierName = ws["D" + str(row)].value
        quantity = float(ws["E" + str(row)].value)
        price = float(ws["F" + str(row)].value)

        c=c+1
        supplier= Supplier.objects.filter(name__contains=supplierName).first()
        article=Article.objects.filter(name__contains=articleName).filter(supplier__name=supplierName).first()
        if article and supplier:
            arrivoFood=ArrivoFood()
            arrivoFood.incomingdate=date
            arrivoFood.article=article
            arrivoFood.quantity=quantity
            arrivoFood.price=price
            arrivoFood.conforme=True
            arrivoFood.scadenza=date.today()
            arrivoFood.save()
            print(arrivoFood.incomingdate,arrivoFood.article.name,arrivoFood.quantity, arrivoFood.price)

def setprice():
    arrivi= ArrivoFood.objects.all()
    for aa in arrivi:


        arr= aa.article
        
        article=Article.objects.filter(name__contains=arr.name).first()
        print(arr.name)
        article.prezzo_unita_arrivo=aa.price
        article.save()









"""


def import_categories():
    wb = load_workbook('/home/mate/Documenti/food_category.xlsx')
    ws = wb.active
    for row in range(2, 22):
        name = ws["B" + str(row)].value
        category = Category()
        category.name = name
        category.save()

    categories = Category.objects.all()
    print(categories)


def import_verdurefreche():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_verfresche.xlsx')
    ws = wb.active
    for row in range(2, 110):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=1)
        ingredient.category = cat

        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals

        print(ingredient.name + ingredient.category.name)
        ingredient.save()


def import_fruttafreca():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_frutta.xlsx')
    ws = wb.active
    for row in range(2, 53):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=2)
        ingredient.category = cat

        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()

        print(ingredient.name + ingredient.category.name)
        ingredient.save()


def import_carnefreca():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_carni_fresceh.xlsx')
    ws = wb.active
    for row in range(2, 77):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=3)
        ingredient.category = cat

        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()

        print(ingredient.name + ingredient.category.name)


def import_latt():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_latt.xlsx')
    ws = wb.active
    for row in range(2, 75):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=4)
        ingredient.category = cat

        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()

        print(ingredient.name + ingredient.category.name)


def import_grassi():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_grassi.xlsx')
    ws = wb.active
    for row in range(2, 10):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=5)
        ingredient.category = cat

        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()

        print(ingredient.name + ingredient.category.name)


def import_cereali():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_cereali.xlsx')
    ws = wb.active
    for row in range(2, 87):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=6)
        ingredient.category = cat

        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()

        print(ingredient.name + ingredient.category.name)


def import_vercong():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_vercong.xlsx')
    ws = wb.active
    for row in range(2, 16):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=8)
        ingredient.category = cat
        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()

        print(ingredient.name + ingredient.category.name)


def import_vercons():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_valvercons.xlsx')
    ws = wb.active
    for row in range(2, 26):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=9)
        ingredient.category = cat
        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()

        print(ingredient.name + ingredient.category.name)


def import_proddolc():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_proddolc.xlsx')
    ws = wb.active
    for row in range(2, 70):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=12)
        ingredient.category = cat
        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()


def import_ittfrechi():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_itticifreschi.xlsx')
    ws = wb.active
    for row in range(2, 84):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=16)
        ingredient.category = cat
        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()


def import_frutcons():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_fruttcons.xlsx')
    ws = wb.active
    for row in range(2, 27):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=20)
        ingredient.category = cat
        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()


def import_carnicons():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_carnicons.xlsx')
    ws = wb.active
    for row in range(2, 43):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=20)
        ingredient.category = cat
        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()


def import_spezie():
    wb = load_workbook('//home/matte/Documents/mirdb/food_ingredient_val_spezie.xlsx')
    ws = wb.active
    for row in range(2, 40):
        name = ws["B" + str(row)].value
        ingredient = Ingredient()
        ingredient.name = name
        cat = Category.objects.get(id=19)
        ingredient.category = cat
        nutritionals = Nutritionals()
        nutritionals.fat = ws["F" + str(row)].value
        nutritionals.calories = ws["c" + str(row)].value
        nutritionals.proteins = ws["E" + str(row)].value
        nutritionals.carbohydrates = ws["D" + str(row)].value
        nutritionals.starch = ws["G" + str(row)].value
        nutritionals.save()
        ingredient.nutritional = nutritionals
        ingredient.save()
